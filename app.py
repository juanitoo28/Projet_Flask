from flask import Flask, render_template, request, redirect, url_for, send_file, jsonify, make_response
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
import os
import io

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads/'  # chemin d'accès pour les fichiers uploadés

# Vérifie si le dossier d'upload existe, sinon le crée
if not os.path.exists(app.config['UPLOAD_FOLDER']):
    os.makedirs(app.config['UPLOAD_FOLDER'])

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload/', methods=['GET', 'POST'])
def upload():
    if request.method == 'POST':
        if 'excel_file' in request.files:
            file = request.files['excel_file']
            filename = secure_filename(file.filename)
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(file_path)
            response = make_response(redirect('/'))
            response.set_cookie('file_path', file_path)
            return response
        else:
            return 'No file part', 400
    else:
        return 'Method Not Allowed', 405

@app.route('/download/')
def download():
    file_path = request.cookies.get('file_path')
    if file_path and os.path.exists(file_path):
        return send_file(file_path, as_attachment=True)
    else:
        return 'File Not Found', 404

def add_sheet(file_path):
    wb = load_workbook(filename=file_path)
    ws = wb.create_sheet("Nouvelle feuille")
    ws.cell(row=1, column=1, value="Exemple de contenu")
    wb.save(file_path)

@app.route('/add_sheet/', methods=['POST'])
def add_sheet_view():
    file_path = request.cookies.get('file_path')
    if file_path and os.path.exists(file_path):
        add_sheet(file_path)
        return 'Sheet Added', 200
    else:
        return 'File Not Found', 404

@app.route('/excel_preview/', methods=['POST'])
def excel_preview():
    if 'excel_file' in request.files:
        file = request.files['excel_file']
        wb = load_workbook(io.BytesIO(file.read()))
        ws = wb.active

        data = []

        for row in ws.iter_rows(values_only=True):
            data.append(list(row))

        return jsonify({"data": data})
    else:
        return jsonify({"error": "Invalid request method"})
    



if __name__ == '__main__':
    app.run(debug=False)
